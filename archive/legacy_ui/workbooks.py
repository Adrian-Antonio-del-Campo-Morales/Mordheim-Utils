"""Persist candidates, enemies, and simulation results in Excel workbooks."""

from __future__ import annotations

from datetime import datetime
import json
import math
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .candidate_catalog import (
    GENERAL_SKILL_DESCRIPTIONS,
    ITEM_TO_OPTION,
    armour_descriptions,
    load_bands,
    weapon_descriptions,
)
from .rules import HOUSE_RULES, POISON_DESCRIPTIONS, PREPARATION_DESCRIPTIONS


FORMAT_VERSION = 1
CATALOG_SCHEMA_VERSION = 1
FORMAT_MARKER = "MORDHEIM_COMBAT_LAB_WORKBOOK_V1"
DATA_SHEET = "_CombatLab"
SUMMARY_SHEET = "Candidate"
ENEMIES_SHEET = "Enemies"
RESULTS_INDEX_SHEET = "Results index"
RESULT_SHEET_PREFIX = "Result · "

NAVY = "17243A"
TEAL = "287D7A"
GOLD = "E9C46A"
PALE = "EEF4F3"
WHITE = "FFFFFF"
GREY = "667085"
THIN_GREY = Side(style="thin", color="D0D5DD")


class CandidateWorkbookError(ValueError):
    pass


_OPTION_IDS = {}
for _item_id, _option in ITEM_TO_OPTION.items():
    _OPTION_IDS.setdefault(_option, _item_id)


def _stable_id(value, prefix: str) -> str:
    if value in (None, "", "None", "No Armour", "No Poison", "Normal"):
        return "none"
    if value in _OPTION_IDS:
        return _OPTION_IDS[value]
    normalized = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", normalized.casefold()).strip("-")
    return f"{prefix}-{slug}"


def _config_ids(config: dict) -> dict:
    """Language-independent identity layer stored alongside display values."""
    return {
        "band_id": config.get("candidate_band_id") or config.get("enemy_band_id") or None,
        "profile_id": config.get("candidate_profile_id") or config.get("enemy_profile_id") or None,
        "main_weapon_id": _stable_id(config.get("main_weapon"), "weapon"),
        "off_hand_id": _stable_id(config.get("off_hand"), "offhand"),
        "armor_id": _stable_id(config.get("armor"), "armor"),
        "main_material_id": _stable_id(config.get("main_weapon_material"), "material"),
        "offhand_material_id": _stable_id(config.get("offhand_material"), "material"),
        "main_poison_id": _stable_id(config.get("main_poison"), "poison"),
        "offhand_poison_id": _stable_id(config.get("offhand_poison"), "poison"),
        "skill_ids": [_stable_id(value, "skill") for value in config.get("skills", ())],
        "preparation_ids": [_stable_id(value, "preparation") for value in config.get("preparations", ())],
    }


def _configured_equipment(config: dict) -> list[str]:
    result = []
    if config.get("has_helmet"):
        result.append("Helmet")
    if config.get("has_luck_amulet"):
        result.append("Lucky charm")
    if config.get("has_sea_dragon_cloak"):
        result.append("Sea Dragon cloak")
    result.extend(config.get("preparations", ()))
    for poison in (
        config.get("main_poison", "No Poison"),
        config.get("offhand_poison", "No Poison"),
    ):
        if poison != "No Poison" and poison not in result:
            result.append(poison)
    return result


def _section(ws, row: int, title: str, end_column: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = ws.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(color=WHITE, bold=True, size=11)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _label_value(ws, row: int, label: str, value, column: int = 1) -> None:
    label_cell = ws.cell(row, column, label)
    value_cell = ws.cell(row, column + 1, "—" if value in (None, "") else value)
    label_cell.font = Font(bold=True, color=NAVY)
    label_cell.fill = PatternFill("solid", fgColor=PALE)
    label_cell.border = value_cell.border = Border(bottom=THIN_GREY)
    label_cell.alignment = Alignment(wrap_text=True, vertical="top")
    value_cell.alignment = Alignment(wrap_text=True, vertical="top")
    text = "—" if value in (None, "") else str(value)
    approximate_width = 28 if column >= 5 else 34
    value_lines = sum(max(1, math.ceil(len(part) / approximate_width)) for part in text.splitlines())
    label_width = 20 if column == 1 else 17
    label_lines = max(1, math.ceil(len(label) / label_width))
    lines = max(value_lines, label_lines)
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 15 * lines)


def _attributes(ws, row: int, config: dict) -> int:
    """Draw the attribute strip with six identical cells."""
    stats = ("WS", "S", "T", "W", "I", "A")
    for column, stat in enumerate(stats, 1):
        header = ws.cell(row, column, stat)
        header.fill = PatternFill("solid", fgColor=GOLD)
        header.font = Font(bold=True, color=NAVY)
        header.alignment = Alignment(horizontal="center", vertical="center")
        value = ws.cell(row + 1, column, config.get(stat, 0))
        value.font = Font(bold=True, size=12)
        value.alignment = Alignment(horizontal="center", vertical="center")
        value.border = Border(bottom=THIN_GREY)
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 22
    return row + 3


def _build_summary(workbook, payload: dict) -> None:
    ws = workbook.create_sheet(SUMMARY_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.merge_cells("A1:H2")
    title = ws["A1"]
    title.value = "MORDHEIM COMBAT LAB · SIMULATION PROFILE"
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(color=WHITE, bold=True, size=18)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws["A3"] = f"Updated: {datetime.now():%d/%m/%Y %H:%M}"
    ws["A3"].font = Font(italic=True, color=GREY)

    config = payload["config"]
    metadata = payload.get("candidate", {})
    row = _section(ws, 5, "CANDIDATE")
    _label_value(ws, row, "Name", metadata.get("name", "Candidate"), 1)
    _label_value(ws, row, "warband", metadata.get("band_name", "Free selection"), 5)
    row += 1
    _label_value(ws, row, "warrior", metadata.get("profile_name", "Free profile"), 1)
    _label_value(ws, row, "Tipo", metadata.get("profile_type", "—"), 5)
    row += 2

    selected_house_rules = [
        HOUSE_RULES[key]
        for key, enabled in (payload.get("house_rules") or {}).items()
        if enabled and key in HOUSE_RULES
    ]
    if selected_house_rules:
        row = _section(ws, row, "ACTIVE HOUSE RULES")
        for rule in selected_house_rules:
            _label_value(ws, row, rule["name"], rule["description"], 1)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
        row += 1

    row = _section(ws, row, "ATTRIBUTES")
    row = _attributes(ws, row, config)

    row = _section(ws, row, "EQUIPMENT AND SKILLS")
    _label_value(ws, row, "Main weapon", config.get("main_weapon"), 1)
    _label_value(ws, row, "Material", config.get("main_weapon_material"), 5)
    row += 1
    _label_value(ws, row, "Off-hand weapon", config.get("off_hand"), 1)
    _label_value(ws, row, "Material", config.get("offhand_material"), 5)
    row += 1
    _label_value(ws, row, "armour", config.get("armor"), 1)
    _label_value(ws, row, "Equipment", ", ".join(_configured_equipment(config)) or "None", 5)
    row += 1
    _label_value(ws, row, "Skills", ", ".join(config.get("skills", ())) or "None", 1)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    row += 2

    row = _section(ws, row, "QUICK REFERENCE")
    weapon_notes = weapon_descriptions()
    armour_notes = armour_descriptions()
    skill_notes = dict(GENERAL_SKILL_DESCRIPTIONS)
    for band in load_bands():
        skill_notes.update((skill.name, skill.description) for skill in band.skills)
    quick_rows = []
    for label, item, descriptions in (
        ("Main weapon", config.get("main_weapon"), weapon_notes),
        ("Off-hand weapon", config.get("off_hand"), weapon_notes),
        ("armour", config.get("armor"), armour_notes),
    ):
        description = descriptions.get(item, "")
        if item and item not in ("None", "No Armour") and description:
            quick_rows.append((f"{label}: {item}", description))
    equipment_notes = {
        **armour_notes,
        **PREPARATION_DESCRIPTIONS,
        **POISON_DESCRIPTIONS,
    }
    quick_rows.extend(
        (f"Equipment: {item}", equipment_notes.get(item, ""))
        for item in _configured_equipment(config)
        if equipment_notes.get(item)
    )
    quick_rows.extend(
        (f"Skill: {skill}", skill_notes.get(skill, "without description canonical breve."))
        for skill in config.get("skills", ())
    )
    if not quick_rows:
        quick_rows.append(("Equipment and skills", "There are no additional short rules to consult."))
    for label, description in quick_rows:
        _label_value(ws, row, label, description, 1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = max(
            ws.row_dimensions[row].height or 15,
            15 * max(1, math.ceil(len(description) / 105)),
        )
        row += 1
    row += 1

    source_rules = metadata.get("rules") or ()
    fixed = metadata.get("fixed_equipment") or ()
    restrictions = metadata.get("restrictions") or ()
    if source_rules or fixed or restrictions:
        row = _section(ws, row, "PROFILE REFERENCE")
        for label, values in (("Fixed equipment", fixed), ("Restrictions", restrictions), ("Rules", source_rules)):
            _label_value(ws, row, label, "\n".join(f"• {value}" for value in values) or "—", 1)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            characters = sum(len(str(value)) for value in values)
            wrapped_lines = max(len(values), math.ceil(characters / 105))
            ws.row_dimensions[row].height = max(22, 15 * max(1, wrapped_lines))
            row += 1
        row += 1

    widths = {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 13, "H": 13}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.print_title_rows = "1:3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _enemy_quick_rows(config):
    notes = weapon_descriptions()
    equipment_notes = {
        **armour_descriptions(),
        **PREPARATION_DESCRIPTIONS,
        **POISON_DESCRIPTIONS,
    }
    skill_notes = dict(GENERAL_SKILL_DESCRIPTIONS)
    for band in load_bands():
        skill_notes.update((skill.name, skill.description) for skill in band.skills)
    rows = []
    for label, key in (("Main weapon", "main_weapon"), ("Off-hand weapon", "off_hand")):
        item = config.get(key)
        if item and item != "None" and notes.get(item):
            rows.append((f"{label}: {item}", notes[item]))
    rows.extend(
        (f"Equipment: {item}", equipment_notes.get(item, ""))
        for item in _configured_equipment(config)
        if equipment_notes.get(item)
    )
    rows.extend(
        (f"Skill: {skill}", skill_notes.get(skill, "without description canonical breve."))
        for skill in config.get("skills", ())
    )
    return rows


def _build_enemies(workbook, payload):
    ws = workbook.create_sheet(ENEMIES_SHEET, 1)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "MORDHEIM COMBAT LAB · ENEMIES"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    data = payload.get("enemies") or {}
    mode = data.get("mode", "sample")
    row = _section(ws, 4, "CONFIGURATION")
    _label_value(ws, row, "Mode", "Random sample" if mode == "sample" else "Manual profiles", 1)
    _label_value(ws, row, "Level", data.get("level", 0), 5)
    row += 2
    if mode == "sample":
        row = _section(ws, row, "RANDOM GENERATION")
        _label_value(ws, row, "Difficulties", ", ".join(data.get("difficulties", ())) or "None", 1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        row += 1
        _label_value(
            ws, row, "Funcionamiento",
            "the simulator genera in each combat a profile ponderado, equipment legal and the improvements indicadas by the level.", 1,
        )
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    else:
        profiles = data.get("profiles") or ()
        for index, config in enumerate(profiles, 1):
            row = _section(ws, row, f"ENEMY {index} · {config.get('enemy_name', f'Enemy {index}')}")
            _label_value(ws, row, "name", config.get("enemy_name", f"Enemy {index}"), 1)
            profile = find_profile_for_workbook(config)
            _label_value(ws, row, "profile", profile, 5)
            row += 1
            row = _section(ws, row, "ATTRIBUTES")
            row = _attributes(ws, row, config)
            _label_value(ws, row, "Main weapon", config.get("main_weapon"), 1)
            _label_value(ws, row, "Off-hand weapon", config.get("off_hand"), 5)
            row += 1
            _label_value(ws, row, "armour", config.get("armor"), 1)
            _label_value(ws, row, "Equipment", ", ".join(_configured_equipment(config)) or "None", 5)
            row += 1
            _label_value(ws, row, "Skills", ", ".join(config.get("skills", ())) or "None", 1)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            for label, description in _enemy_quick_rows(config):
                _label_value(ws, row, label, description, 1)
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
                row += 1
            row += 1
    for column, width in {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 13, "H": 13}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def find_profile_for_workbook(config):
    from .candidate_catalog import find_profile
    profile = find_profile(
        config.get("enemy_band_id", ""), config.get("enemy_profile_id", "")
    )
    return f"{profile.band_name} · {profile.name}" if profile else "Free selection"


def _result_sheet_name(title: str, used: set[str]) -> str:
    base = f"{RESULT_SHEET_PREFIX}{title}"[:31]
    name = base
    suffix = 2
    while name in used:
        tail = f" {suffix}"
        name = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(name)
    return name


def _build_result_sheet(workbook, result: dict, sheet_name: str) -> None:
    ws = workbook.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    headers = tuple(result.get("headers") or ())
    end_column = max(6, len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    ws["A1"] = f"MORDHEIM COMBAT LAB · {result.get('title', 'results').upper()}"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    _label_value(ws, 4, "Fecha", result.get("generated_at", "—"), 1)
    _label_value(ws, 4, "Iteraciones", result.get("iterations", 0), 5)
    _label_value(ws, 5, "opponent", result.get("opponent", "—"), 1)
    _label_value(ws, 5, "View", result.get("view", "—"), 5)
    for column, header in enumerate(headers, 1):
        cell = ws.cell(7, column, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, values in enumerate(result.get("rows") or (), 8):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.00"
    for column in range(1, max(1, len(headers)) + 1):
        letter = get_column_letter(column)
        longest = max(
            [len(str(ws.cell(row, column).value or "")) for row in range(7, ws.max_row + 1)]
            or [12]
        )
        ws.column_dimensions[letter].width = min(42, max(13, longest + 2))
    ws.auto_filter.ref = f"A7:{ws.cell(max(7, ws.max_row), max(1, len(headers))).coordinate}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _build_results(workbook, results) -> None:
    ws = workbook.create_sheet(RESULTS_INDEX_SHEET)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F2")
    ws["A1"] = "SAVED RESULTS"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A4"] = "Saved simulations will appear here and in their own worksheets."
    ws.merge_cells("A4:F4")
    headers = ("Date", "Simulation", "Iterations", "Opponent", "View", "Sheet")
    for column, value in enumerate(headers, 1):
        cell = ws.cell(6, column, value)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(color=WHITE, bold=True)
    for column, width in enumerate((20, 24, 14, 28, 16, 24), 1):
        ws.column_dimensions[chr(64 + column)].width = width
    ws.freeze_panes = "A7"
    used = set(workbook.sheetnames)
    for row, result in enumerate(results or (), 7):
        sheet_name = _result_sheet_name(str(result.get("title", "Simulation")), used)
        _build_result_sheet(workbook, result, sheet_name)
        values = (
            result.get("generated_at", "—"), result.get("title", "Simulation"),
            result.get("iterations", 0), result.get("opponent", "—"),
            result.get("view", "—"), sheet_name,
        )
        for column, value in enumerate(values, 1):
            ws.cell(row, column, value)
        ws.cell(row, 6).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row, 6).style = "Hyperlink"


def save_candidate_workbook(path, payload: dict) -> Path:
    destination = Path(path)
    workbook = Workbook()
    del workbook["Sheet"]
    payload = {
        **payload,
        "format_version": FORMAT_VERSION,
        "catalog_schema_version": CATALOG_SCHEMA_VERSION,
        "catalog_categories": payload.get(
            "catalog_categories", ["core", "1a", "1b", "1c", "trollheim"]
        ),
        "locale": payload.get("locale", "en"),
        "config_ids": _config_ids(payload.get("config", {})),
        "enemy_config_ids": [
            _config_ids(config)
            for config in (payload.get("enemies", {}).get("profiles", ()))
        ],
        "results": list(payload.get("results", ())),
    }
    _build_summary(workbook, payload)
    _build_enemies(workbook, payload)
    _build_results(workbook, payload.get("results", ()))
    data = workbook.create_sheet(DATA_SHEET)
    data.sheet_state = "veryHidden"
    data["A1"] = FORMAT_MARKER
    data["A2"] = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    data["A3"] = FORMAT_VERSION
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def load_candidate_workbook(path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if DATA_SHEET not in workbook.sheetnames:
        raise CandidateWorkbookError("The workbook does not contain a Mordheim Combat Lab profile.")
    data = workbook[DATA_SHEET]
    if data["A1"].value != FORMAT_MARKER or data["A3"].value != FORMAT_VERSION:
        raise CandidateWorkbookError("This Combat Lab workbook version is not compatible.")
    try:
        payload = json.loads(data["A2"].value)
    except (TypeError, json.JSONDecodeError) as exc:
        raise CandidateWorkbookError("The profile's internal data is corrupted.") from exc
    if isinstance(payload, dict) and "catalog_categories" not in payload:
        legacy_category = str(payload.get("catalog_category", "all")).casefold()
        payload["catalog_categories"] = (
            ["core", "1a", "1b", "1c", "trollheim"]
            if legacy_category == "all"
            else [legacy_category]
        )
    if (
        not isinstance(payload, dict)
        or payload.get("format_version") != FORMAT_VERSION
        or payload.get("catalog_schema_version") != CATALOG_SCHEMA_VERSION
        or not isinstance(payload.get("catalog_categories"), list)
        or not isinstance(payload.get("config"), dict)
        or not isinstance(payload.get("config_ids"), dict)
    ):
        raise CandidateWorkbookError("The profile does not contain a valid candidate.")
    enemies = payload.get("enemies")
    if (
        not isinstance(enemies, dict)
        or enemies.get("mode") not in {"sample", "custom"}
        or not isinstance(enemies.get("profiles"), list)
    ):
        raise CandidateWorkbookError("The workbook does not contain a valid enemy configuration.")
    results = payload.get("results")
    if not isinstance(results, list) or any(
        not isinstance(result, dict)
        or result.get("format_version") != FORMAT_VERSION
        or result.get("target") not in {"combos", "weapons", "equipment"}
        for result in results
    ):
        raise CandidateWorkbookError("The workbook does not contain results in the current format.")
    return payload
