"""persistence.workbooks: responsibility extracted without altering the rules."""
from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from datetime import timezone
import json as json
from mordheim_combat_lab.application.settings import DuelExecutionSettings
from mordheim_core.models import Characteristics
from mordheim_core.models import DuelResult
from mordheim_core.models import FighterBuild
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path
from zipfile import BadZipFile


WORKBOOK_MARKER = "MORDHEIM_COMBAT_LAB_UI"


WORKBOOK_VERSION = 1


DATA_SHEET = "_MordheimCombatLab"


SUMMARY_SHEET = "Configuration"


RESULTS_SHEET = "Last result"


class CombatLabWorkbookError(ValueError):
    """Raised when a workbook does not satisfy the current UI schema."""


def _build_payload(build: FighterBuild) -> dict:
    payload = asdict(build)
    payload["characteristics"] = asdict(build.characteristics) if build.characteristics else None
    return payload


def _build_from_payload(payload: dict) -> FighterBuild:
    try:
        characteristics = payload.get("characteristics")
        return FighterBuild(
            ruleset=str(payload["ruleset"]),
            characteristics=Characteristics(**characteristics) if characteristics else None,
            band_id=payload.get("band_id"), profile_id=payload.get("profile_id"),
            main_weapon_id=str(payload.get("main_weapon_id") or "weapon.dagger"),
            off_hand_id=payload.get("off_hand_id"), armour_id=str(payload.get("armour_id") or "armour.no-armour"),
            defence_ids=tuple(payload.get("defence_ids") or ()),
            main_material_id=str(payload.get("main_material_id") or "material.normal"),
            off_material_id=str(payload.get("off_material_id") or "material.normal"),
            skill_ids=tuple(payload.get("skill_ids") or ()), preparation_ids=tuple(payload.get("preparation_ids") or ()),
            special_rule_ids=tuple(payload.get("special_rule_ids") or ()),
            energy_focus_attacks=int(payload.get("energy_focus_attacks") or 0),
            variant_ids=tuple(payload.get("variant_ids") or ()),
            main_poison_id=payload.get("main_poison_id"), off_poison_id=payload.get("off_poison_id"),
            trait_overrides=dict(payload.get("trait_overrides") or {}), collection=str(payload.get("collection") or "mordheim"),
        )
    except (KeyError, TypeError, ValueError) as exc:
        raise CombatLabWorkbookError(f"Invalid fighter build payload: {exc}") from exc


def save_workbook(path, candidate: FighterBuild, enemy: FighterBuild, settings: DuelExecutionSettings, result: DuelResult | None = None) -> Path:
    """Write an English, stable-ID workbook for the active new UI."""
    destination = Path(path)
    payload = {
        "marker": WORKBOOK_MARKER,
        "version": WORKBOOK_VERSION,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "candidate": _build_payload(candidate),
        "enemy": _build_payload(enemy),
        "execution": asdict(settings),
        "result": asdict(result) if result else None,
    }
    workbook = Workbook()
    summary = workbook.active
    summary.title = SUMMARY_SHEET
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "MORDHEIM COMBAT LAB · UI CONFIGURATION"
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].fill = PatternFill("solid", fgColor="17243A")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.append([])
    for row in (("Saved (UTC)", payload["saved_at"]), ("Candidate", _display_build(candidate)), ("Enemy", _display_build(enemy)), ("Simulations", settings.simulations), ("Seed", settings.seed), ("Batch size", settings.batch_size), ("Maximum rounds", settings.maximum_rounds)):
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 80

    if result:
        results = workbook.create_sheet(RESULTS_SHEET)
        results.append(("Candidate wins", result.first_wins))
        results.append(("Enemy wins", result.second_wins))
        results.append(("Unresolved", result.unresolved))
        results.append(("Simulations", result.simulations))
        results.append(("Candidate win rate", result.first_win_rate / 100))
        results.append(("Enemy win rate", result.second_win_rate / 100))
        results.append(("Unresolved rate", result.unresolved_rate / 100))
        for row in range(5, 8):
            results.cell(row, 2).number_format = "0.00%"
        results.column_dimensions["A"].width = 24
        results.column_dimensions["B"].width = 18

    data = workbook.create_sheet(DATA_SHEET)
    data.sheet_state = "hidden"
    data["A1"] = WORKBOOK_MARKER
    data["A2"] = WORKBOOK_VERSION
    data["A3"] = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def load_ui_workbook(path) -> tuple[FighterBuild, FighterBuild, DuelExecutionSettings, DuelResult | None]:
    """Load a workbook produced by :func:`save_workbook`."""
    workbook = None
    try:
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        data = workbook[DATA_SHEET]
        if data["A1"].value != WORKBOOK_MARKER or data["A2"].value != WORKBOOK_VERSION:
            raise CombatLabWorkbookError("Workbook is not a supported Mordheim Combat Lab UI workbook.")
        payload = json.loads(str(data["A3"].value))
        candidate = _build_from_payload(dict(payload["candidate"]))
        enemy = _build_from_payload(dict(payload["enemy"]))
        execution = DuelExecutionSettings(**dict(payload["execution"]))
        result_payload = payload.get("result")
        result = DuelResult(**result_payload) if result_payload else None
        return candidate, enemy, execution, result
    except CombatLabWorkbookError:
        raise
    except (BadZipFile, InvalidFileException, KeyError, OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise CombatLabWorkbookError(f"Could not load workbook: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _display_build(build: FighterBuild) -> str:
    if build.band_id and build.profile_id:
        return f"{build.collection}/{build.band_id}/{build.profile_id}"
    return "Free selection"
